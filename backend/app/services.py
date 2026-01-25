import base64
import math
import os
import struct
import zipfile
from datetime import datetime
from typing import List, Optional, Tuple

from sqlalchemy.orm import Session

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None

try:
    import ezdxf
    from ezdxf import units as dxf_units
except ImportError:  # pragma: no cover - optional dependency
    ezdxf = None
    dxf_units = None

from app import models
from app.config import DEFAULT_FLOOR_HEIGHT, STORAGE_DIR


def _safe_text(value: Optional[str], fallback: str = "") -> str:
    return value if value is not None else fallback


def _safe_number(value: Optional[str], fallback: float) -> float:
    if value is None:
        return fallback
    digits = "".join([char for char in str(value) if (char.isdigit() or char == ".")])
    if not digits:
        return fallback
    try:
        return float(digits)
    except ValueError:
        return fallback


def _ensure_storage(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _write_file(path: str, content: str) -> None:
    with open(path, "w", encoding="utf-8") as handle:
        handle.write(content)


def _write_binary(path: str, content: bytes) -> None:
    with open(path, "wb") as handle:
        handle.write(content)


def _compute_massing(project: models.Project) -> dict:
    floors = _safe_number(project.floors, 8)
    gfa = _safe_number(project.gfa, 18000)
    floor_area = gfa / max(floors, 1)

    building_type = (project.building_type or "").lower()
    region = (project.region or "").lower()

    ratio = 1.2
    module = 7.5
    default_core_ratio = 0.24

    if "residential" in building_type or "apartment" in building_type:
        ratio = 1.55
        module = 6.0
        default_core_ratio = 0.18
    elif "hospital" in building_type or "health" in building_type:
        ratio = 1.25
        module = 8.0
        default_core_ratio = 0.28
    elif "education" in building_type:
        ratio = 1.4
        module = 7.2
        default_core_ratio = 0.22
    elif "hospitality" in building_type or "hotel" in building_type:
        ratio = 1.3
        module = 6.8
        default_core_ratio = 0.2
    elif "mixed" in building_type:
        ratio = 1.28
        module = 7.2
        default_core_ratio = 0.24

    if "egypt" in region:
        module *= 0.95

    width = max((floor_area ** 0.5) / ratio, 18)
    depth = max((floor_area ** 0.5) * ratio, 14)

    core_ratio = _safe_number(project.core_ratio, default_core_ratio * 100) / 100
    core_ratio = min(max(core_ratio, 0.12), 0.35)

    grid_x = max(int(width / module) + 1, 3)
    grid_y = max(int(depth / module) + 1, 3)

    return {
        "width": width,
        "depth": depth,
        "height": max(floors * DEFAULT_FLOOR_HEIGHT, 12),
        "floors": floors,
        "floor_area": floor_area,
        "core_ratio": core_ratio,
        "module": module,
        "grid_x": grid_x,
        "grid_y": grid_y,
    }


def _build_plan_svg(project: models.Project, massing: Optional[dict] = None) -> str:
    if massing is None:
        massing = _compute_massing(project)
    width = massing["width"]
    depth = massing["depth"]
    core_ratio = massing["core_ratio"]
    grid_x = massing["grid_x"]
    grid_y = massing["grid_y"]

    core_width = width * 0.32
    core_area = width * depth * core_ratio
    core_depth = min(core_area / max(core_width, 1), depth * 0.6)
    core_x = (width - core_width) / 2
    core_y = (depth - core_depth) / 2
    grid_weight = 0.4
    wall_weight = 1.2
    partition_weight = 0.8

    grid_lines = []
    grid_positions_x = []
    for i in range(grid_x):
        x = 0.5 + i * (width - 1) / (grid_x - 1)
        grid_positions_x.append(x)
        grid_lines.append(
            f"<line x1='{x:.2f}' y1='0.5' x2='{x:.2f}' y2='{depth - 0.5:.2f}' "
            f"stroke='#e0d2c2' stroke-width='{grid_weight:.2f}' />"
        )
    grid_positions_y = []
    for j in range(grid_y):
        y = 0.5 + j * (depth - 1) / (grid_y - 1)
        grid_positions_y.append(y)
        grid_lines.append(
            f"<line x1='0.5' y1='{y:.2f}' x2='{width - 0.5:.2f}' y2='{y:.2f}' "
            f"stroke='#e0d2c2' stroke-width='{grid_weight:.2f}' />"
        )

    columns = []
    for i in range(grid_x):
        x = 0.5 + i * (width - 1) / (grid_x - 1)
        for j in range(grid_y):
            y = 0.5 + j * (depth - 1) / (grid_y - 1)
            columns.append(
                f"<circle cx='{x:.2f}' cy='{y:.2f}' r='0.35' fill='#c9b9aa' />"
            )

    return (
        f"<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 {width:.2f} {depth:.2f}' "
        f"data-grid-x='{','.join(f'{value:.2f}' for value in grid_positions_x)}' "
        f"data-grid-y='{','.join(f'{value:.2f}' for value in grid_positions_y)}' "
        "shape-rendering='geometricPrecision'>"
        "<style>[stroke]{vector-effect:non-scaling-stroke;stroke-linecap:square;"
        "stroke-linejoin:miter;}</style>"
        f"<rect x='0.5' y='0.5' width='{width - 1:.2f}' height='{depth - 1:.2f}' "
        f"fill='#fbf6ef' stroke='#c7b8a8' stroke-width='{wall_weight:.2f}' />"
        f"<rect x='{core_x:.2f}' y='{core_y:.2f}' width='{core_width:.2f}' height='{core_depth:.2f}' "
        f"fill='#e6d9cb' stroke='#b8a897' stroke-width='{partition_weight:.2f}' />"
        + "".join(grid_lines)
        + "".join(columns)
        + "</svg>"
    )


def _coerce_float_list(values: Optional[list]) -> list[float]:
    result = []
    for value in values or []:
        try:
            result.append(float(value))
        except (TypeError, ValueError):
            continue
    return result


def _axis_label(index: int, axis: str = "x") -> str:
    if axis == "x":
        label = ""
        value = index
        while True:
            label = chr(ord("A") + (value % 26)) + label
            value = value // 26 - 1
            if value < 0:
                break
        return label
    return str(index + 1)


def _format_distance(value: float) -> str:
    return f"{value:.1f} m"


def _normalize_grid(grid: List[float], total: float, count: int) -> List[float]:
    if grid:
        return grid
    count = max(count, 2)
    step = total / (count - 1)
    return [i * step for i in range(count)]


def _clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(value, upper))


def _build_arch_plan_svg(
    plan: dict,
    structural: Optional[dict] = None,
    mep: Optional[dict] = None,
    massing: Optional[dict] = None,
) -> str:
    plan = plan or {}
    structural = structural or {}
    mep = mep or {}
    massing = massing or {}

    width = _safe_number(plan.get("width"), _safe_number(massing.get("width"), 40))
    depth = _safe_number(plan.get("depth"), _safe_number(massing.get("depth"), 30))
    width = max(width, 10)
    depth = max(depth, 10)

    grid_x = _coerce_float_list(plan.get("grid_x"))
    grid_y = _coerce_float_list(plan.get("grid_y"))
    grid_count_x = int(_safe_number(massing.get("grid_x"), 4)) if massing else 4
    grid_count_y = int(_safe_number(massing.get("grid_y"), 4)) if massing else 4
    grid_x = _normalize_grid(grid_x, width, grid_count_x)
    grid_y = _normalize_grid(grid_y, depth, grid_count_y)
    grid_x = sorted(set(grid_x))
    grid_y = sorted(set(grid_y))
    if len(grid_x) < 2:
        grid_x = [0.0, width]
    if len(grid_y) < 2:
        grid_y = [0.0, depth]

    core = plan.get("core") or {}
    if core:
        core_pos = core.get("position") or [width * 0.35, depth * 0.35]
        try:
            core_x = float(core_pos[0])
            core_y = float(core_pos[1])
        except (TypeError, IndexError, ValueError):
            core_x = width * 0.35
            core_y = depth * 0.35
        core_w = _safe_number(core.get("width"), width * 0.3)
        core_d = _safe_number(core.get("depth"), depth * 0.3)
        core_elevators = int(_safe_number(core.get("elevators"), 2))
        core_stairs = int(_safe_number(core.get("stairs"), 2))
    else:
        core_ratio = _safe_number(massing.get("core_ratio"), 0.2)
        core_w = width * 0.32
        core_area = width * depth * core_ratio
        core_d = min(core_area / max(core_w, 1), depth * 0.6)
        core_x = (width - core_w) / 2
        core_y = (depth - core_d) / 2
        core_elevators = 2
        core_stairs = 2

    core_w = min(core_w, width * 0.7)
    core_d = min(core_d, depth * 0.7)
    core_x = _clamp(core_x, 0.5, width - core_w - 0.5)
    core_y = _clamp(core_y, 0.5, depth - core_d - 0.5)

    spaces = plan.get("spaces") or []
    circulation = plan.get("circulation") or []
    has_corridor_space = any(
        (space.get("type") or "").lower() in {"corridor", "circulation"} for space in spaces
    )

    palette = {
        "apartment": "#f2e7d5",
        "corridor": "#ede1d2",
        "core": "#e4d4c2",
        "service": "#f5efe8",
        "retail": "#f0dcc9",
        "lobby": "#efe2d1",
        "meeting_room": "#f3e1cc",
        "open_office": "#f0e7da",
        "private_office": "#eadfce",
        "guest_room": "#f3e5d6",
        "support": "#f6eee5",
        "circulation": "#efe2d6",
        "office": "#f2e9dd",
    }

    margin = min(max(max(width, depth) * 0.12, 4.0), 10.0)
    view_box = f"{-margin:.2f} {-margin:.2f} {width + 2 * margin:.2f} {depth + 2 * margin:.2f}"
    grid_x_attr = ",".join(f"{value:.2f}" for value in grid_x)
    grid_y_attr = ",".join(f"{value:.2f}" for value in grid_y)
    axis_font = max(0.9, margin * 0.24)
    dim_font = max(0.85, margin * 0.22)
    label_font = max(1.0, min(width, depth) * 0.03)
    legend_font = max(0.8, margin * 0.2)
    # Professional CAD line weights (matching AutoCAD/Revit standards)
    # Thinner lines for cleaner, more professional appearance
    wall_weight = 0.7              # External/structural walls - bold but not heavy
    partition_weight = 0.45        # Interior partitions - medium
    axis_weight = 0.3              # Grid axis lines - light
    grid_weight = 0.2              # Background grid - very light
    dim_weight = 0.35              # Dimension lines - thin
    dim_ext_weight = 0.25          # Dimension extension lines - hairline
    dim_tick_weight = 0.35         # Dimension tick marks
    window_weight = 0.35           # Window symbols - medium-light
    door_weight = 0.4              # Door symbols - medium
    detail_weight = 0.3            # Detail elements - light
    struct_weight = 0.5            # Structural elements - medium-bold
    struct_light_weight = 0.35     # Light structural elements
    mep_weight = 0.45              # MEP elements - medium
    mep_light_weight = 0.3         # Light MEP elements
    legend_weight = 0.35           # Legend elements

    elements = [
        f"<svg xmlns='http://www.w3.org/2000/svg' viewBox='{view_box}' "
        "preserveAspectRatio='xMidYMid meet' shape-rendering='geometricPrecision' "
        f"data-grid-x='{grid_x_attr}' data-grid-y='{grid_y_attr}'>",
        "<style>"
        f".axis-label{{font-family:'Space Mono',monospace;font-size:{axis_font:.2f}px;fill:#5d4c3c;font-weight:500;}}"
        f".dim-text{{font-family:'Space Mono',monospace;font-size:{dim_font:.2f}px;fill:#5d4c3c;font-weight:500;}}"
        f".space-label{{font-family:'Sora',sans-serif;font-size:{label_font:.2f}px;fill:#6a594a;font-weight:500;}}"
        f".space-size{{font-family:'Sora',sans-serif;font-size:{label_font * 0.75:.2f}px;fill:#8a7a6b;font-weight:400;}}"
        f".legend-text{{font-family:'Space Mono',monospace;font-size:{legend_font:.2f}px;fill:#5d4c3c;font-weight:500;}}"
        ".layer-axes [stroke],.layer-dims [stroke],.layer-arch [stroke],.layer-struct [stroke],"
        ".layer-mep [stroke],.layer-legend [stroke]{vector-effect:non-scaling-stroke;"
        "stroke-linecap:square;stroke-linejoin:miter;}"
        "</style>",
    ]

    arch_elements: List[str] = []
    axis_elements: List[str] = []
    dim_elements: List[str] = []
    struct_elements: List[str] = []
    mep_elements: List[str] = []
    legend_elements: List[str] = []

    def door_symbol(x: float, y: float, width_s: float, orientation: str) -> str:
        stroke = "#7d6a58"
        width_s = max(0.7, min(width_s, 1.4))
        if orientation == "N":
            return (
                f"<line x1='{x:.2f}' y1='{y:.2f}' x2='{x + width_s:.2f}' y2='{y:.2f}' "
                f"stroke='{stroke}' stroke-width='{door_weight:.2f}' />"
                f"<path d='M{x:.2f} {y:.2f} A{width_s:.2f} {width_s:.2f} 0 0 0 {x:.2f} "
                f"{y - width_s:.2f}' stroke='{stroke}' stroke-width='{door_weight:.2f}' fill='none' />"
            )
        if orientation == "S":
            return (
                f"<line x1='{x:.2f}' y1='{y:.2f}' x2='{x + width_s:.2f}' y2='{y:.2f}' "
                f"stroke='{stroke}' stroke-width='{door_weight:.2f}' />"
                f"<path d='M{x:.2f} {y:.2f} A{width_s:.2f} {width_s:.2f} 0 0 1 {x:.2f} "
                f"{y + width_s:.2f}' stroke='{stroke}' stroke-width='{door_weight:.2f}' fill='none' />"
            )
        if orientation == "E":
            return (
                f"<line x1='{x:.2f}' y1='{y:.2f}' x2='{x:.2f}' y2='{y + width_s:.2f}' "
                f"stroke='{stroke}' stroke-width='{door_weight:.2f}' />"
                f"<path d='M{x:.2f} {y:.2f} A{width_s:.2f} {width_s:.2f} 0 0 1 {x + width_s:.2f} "
                f"{y:.2f}' stroke='{stroke}' stroke-width='{door_weight:.2f}' fill='none' />"
            )
        if orientation == "W":
            return (
                f"<line x1='{x:.2f}' y1='{y:.2f}' x2='{x:.2f}' y2='{y + width_s:.2f}' "
                f"stroke='{stroke}' stroke-width='{door_weight:.2f}' />"
                f"<path d='M{x:.2f} {y:.2f} A{width_s:.2f} {width_s:.2f} 0 0 0 {x - width_s:.2f} "
                f"{y:.2f}' stroke='{stroke}' stroke-width='{door_weight:.2f}' fill='none' />"
            )
        return ""

    def add_windows_vertical(y_start: float, y_end: float, x1: float, x2: float) -> None:
        length = y_end - y_start
        if length <= 1.2:
            return
        count = max(1, int(length / 2.8))
        for i in range(count):
            y = y_start + (i + 1) * length / (count + 1)
            arch_elements.append(
                f"<line x1='{x1:.2f}' y1='{y:.2f}' x2='{x2:.2f}' y2='{y:.2f}' "
                f"stroke='#8aa6c1' stroke-width='{window_weight:.2f}' />"
            )
            arch_elements.append(
                f"<line x1='{x1:.2f}' y1='{y + 0.3:.2f}' x2='{x2:.2f}' y2='{y + 0.3:.2f}' "
                f"stroke='#8aa6c1' stroke-width='{window_weight:.2f}' />"
            )

    def add_windows_horizontal(x_start: float, x_end: float, y1: float, y2: float) -> None:
        length = x_end - x_start
        if length <= 1.2:
            return
        count = max(1, int(length / 2.8))
        for i in range(count):
            x = x_start + (i + 1) * length / (count + 1)
            arch_elements.append(
                f"<line x1='{x:.2f}' y1='{y1:.2f}' x2='{x:.2f}' y2='{y2:.2f}' "
                f"stroke='#8aa6c1' stroke-width='{window_weight:.2f}' />"
            )
            arch_elements.append(
                f"<line x1='{x + 0.3:.2f}' y1='{y1:.2f}' x2='{x + 0.3:.2f}' y2='{y2:.2f}' "
                f"stroke='#8aa6c1' stroke-width='{window_weight:.2f}' />"
            )

    def distribute_points(
        min_x: float,
        min_y: float,
        max_x: float,
        max_y: float,
        count: int,
    ) -> List[Tuple[float, float]]:
        if count <= 0:
            return []
        width_s = max(max_x - min_x, 0.5)
        depth_s = max(max_y - min_y, 0.5)
        cols = max(1, int(math.sqrt(count)))
        rows = int(math.ceil(count / cols))
        step_x = width_s / (cols + 1)
        step_y = depth_s / (rows + 1)
        points = []
        idx = 0
        for r in range(rows):
            for c in range(cols):
                if idx >= count:
                    break
                points.append((min_x + (c + 1) * step_x, min_y + (r + 1) * step_y))
                idx += 1
        return points

    def entry_orientation(min_x: float, min_y: float, max_x: float, max_y: float) -> str:
        core_center_x = core_x + core_w / 2
        core_center_y = core_y + core_d / 2
        center_x = (min_x + max_x) / 2
        center_y = (min_y + max_y) / 2
        dx = core_center_x - center_x
        dy = core_center_y - center_y
        if abs(dx) > abs(dy):
            return "E" if dx > 0 else "W"
        return "S" if dy > 0 else "N"

    def space_label(space_type: str, name: str) -> str:
        if space_type in {"apartment", "guest_room"}:
            digits = "".join([ch for ch in str(name or "") if ch.isdigit()])
            return f"Unit {digits}" if digits else "Unit"
        if space_type == "corridor":
            return "Corridor"
        if space_type == "service":
            return "Service"
        return (name or space_type.replace("_", " ").title())[:18]

    def add_unit_detail(
        min_x: float,
        min_y: float,
        max_x: float,
        max_y: float,
        orientation: str,
    ) -> None:
        width_s = max_x - min_x
        depth_s = max_y - min_y
        if width_s < 4.2 or depth_s < 3.6:
            return
        wet_w = min(2.2, width_s * 0.32)
        wet_d = min(2.4, depth_s * 0.32)
        detail_stroke = "#cdbdaa"
        if orientation in {"E", "W"}:
            if orientation == "W":
                wet_x0 = min_x
                wet_x1 = min_x + wet_w
            else:
                wet_x1 = max_x
                wet_x0 = max_x - wet_w
            wet_y0 = min_y
            wet_y1 = min_y + wet_d
            split_x = min_x + width_s * 0.55
            arch_elements.append(
                f"<rect x='{wet_x0:.2f}' y='{wet_y0:.2f}' width='{wet_x1 - wet_x0:.2f}' "
                f"height='{wet_y1 - wet_y0:.2f}' fill='none' stroke='{detail_stroke}' "
                f"stroke-width='{detail_weight:.2f}' />"
            )
            arch_elements.append(
                f"<line x1='{split_x:.2f}' y1='{min_y + 0.4:.2f}' "
                f"x2='{split_x:.2f}' y2='{max_y - 0.4:.2f}' "
                f"stroke='{detail_stroke}' stroke-width='{detail_weight:.2f}' />"
            )
        else:
            if orientation == "S":
                wet_y1 = max_y
                wet_y0 = max_y - wet_d
            else:
                wet_y0 = min_y
                wet_y1 = min_y + wet_d
            wet_x0 = min_x
            wet_x1 = min_x + wet_w
            split_y = min_y + depth_s * 0.55
            arch_elements.append(
                f"<rect x='{wet_x0:.2f}' y='{wet_y0:.2f}' width='{wet_x1 - wet_x0:.2f}' "
                f"height='{wet_y1 - wet_y0:.2f}' fill='none' stroke='{detail_stroke}' "
                f"stroke-width='{detail_weight:.2f}' />"
            )
            arch_elements.append(
                f"<line x1='{min_x + 0.4:.2f}' y1='{split_y:.2f}' "
                f"x2='{max_x - 0.4:.2f}' y2='{split_y:.2f}' "
                f"stroke='{detail_stroke}' stroke-width='{detail_weight:.2f}' />"
            )

    # Axis grid and labels
    for i, x in enumerate(grid_x):
        axis_elements.append(
            f"<line x1='{x:.2f}' y1='0.5' x2='{x:.2f}' y2='{depth - 0.5:.2f}' "
            f"stroke='#e1d4c4' stroke-width='{grid_weight:.2f}' />"
        )
        label = _axis_label(i, "x")
        label_y = -margin * 0.45
        axis_elements.append(
            f"<circle cx='{x:.2f}' cy='{label_y:.2f}' r='{axis_font * 0.55:.2f}' "
            f"fill='#fff7ef' stroke='#b59d88' stroke-width='{axis_weight:.2f}' />"
        )
        axis_elements.append(
            f"<text class='axis-label' x='{x:.2f}' y='{label_y + axis_font * 0.35:.2f}' "
            "text-anchor='middle'>"
            f"{label}</text>"
        )

    for i, y in enumerate(grid_y):
        axis_elements.append(
            f"<line x1='0.5' y1='{y:.2f}' x2='{width - 0.5:.2f}' y2='{y:.2f}' "
            f"stroke='#e1d4c4' stroke-width='{grid_weight:.2f}' />"
        )
        label = _axis_label(i, "y")
        label_x = -margin * 0.45
        axis_elements.append(
            f"<circle cx='{label_x:.2f}' cy='{y:.2f}' r='{axis_font * 0.55:.2f}' "
            f"fill='#fff7ef' stroke='#b59d88' stroke-width='{axis_weight:.2f}' />"
        )
        axis_elements.append(
            f"<text class='axis-label' x='{label_x:.2f}' y='{y + axis_font * 0.35:.2f}' "
            "text-anchor='middle'>"
            f"{label}</text>"
        )

    # Dimensioning
    tick = max(0.4, margin * 0.12)
    dim_offset = margin * 0.75
    grid_offset = margin * 0.4
    text_offset = max(0.5, margin * 0.18)

    # Overall width
    dim_y = depth + dim_offset
    dim_elements.append(
        f"<line x1='0.5' y1='{dim_y:.2f}' x2='{width - 0.5:.2f}' y2='{dim_y:.2f}' "
        f"stroke='#8d7b6a' stroke-width='{dim_weight:.2f}' />"
    )
    dim_elements.append(
        f"<line x1='0.5' y1='{dim_y - tick:.2f}' x2='0.5' y2='{dim_y + tick:.2f}' "
        f"stroke='#8d7b6a' stroke-width='{dim_tick_weight:.2f}' />"
    )
    dim_elements.append(
        f"<line x1='{width - 0.5:.2f}' y1='{dim_y - tick:.2f}' x2='{width - 0.5:.2f}' "
        f"y2='{dim_y + tick:.2f}' stroke='#8d7b6a' stroke-width='{dim_tick_weight:.2f}' />"
    )
    dim_elements.append(
        f"<line x1='0.5' y1='{depth:.2f}' x2='0.5' y2='{dim_y:.2f}' "
        f"stroke='#b9a392' stroke-width='{dim_ext_weight:.2f}' />"
    )
    dim_elements.append(
        f"<line x1='{width - 0.5:.2f}' y1='{depth:.2f}' x2='{width - 0.5:.2f}' y2='{dim_y:.2f}' "
        f"stroke='#b9a392' stroke-width='{dim_ext_weight:.2f}' />"
    )
    dim_elements.append(
        f"<text class='dim-text' x='{width / 2:.2f}' y='{dim_y - text_offset:.2f}' "
        "text-anchor='middle'>"
        f"{_format_distance(width)}</text>"
    )

    # Overall depth
    dim_x = width + dim_offset
    dim_elements.append(
        f"<line x1='{dim_x:.2f}' y1='0.5' x2='{dim_x:.2f}' y2='{depth - 0.5:.2f}' "
        f"stroke='#8d7b6a' stroke-width='{dim_weight:.2f}' />"
    )
    dim_elements.append(
        f"<line x1='{dim_x - tick:.2f}' y1='0.5' x2='{dim_x + tick:.2f}' y2='0.5' "
        f"stroke='#8d7b6a' stroke-width='{dim_tick_weight:.2f}' />"
    )
    dim_elements.append(
        f"<line x1='{dim_x - tick:.2f}' y1='{depth - 0.5:.2f}' x2='{dim_x + tick:.2f}' "
        f"y2='{depth - 0.5:.2f}' stroke='#8d7b6a' stroke-width='{dim_tick_weight:.2f}' />"
    )
    dim_elements.append(
        f"<line x1='{width:.2f}' y1='0.5' x2='{dim_x:.2f}' y2='0.5' "
        f"stroke='#b9a392' stroke-width='{dim_ext_weight:.2f}' />"
    )
    dim_elements.append(
        f"<line x1='{width:.2f}' y1='{depth - 0.5:.2f}' x2='{dim_x:.2f}' "
        f"y2='{depth - 0.5:.2f}' stroke='#b9a392' stroke-width='{dim_ext_weight:.2f}' />"
    )
    dim_elements.append(
        f"<text class='dim-text' x='{dim_x + text_offset:.2f}' y='{depth / 2:.2f}' "
        "text-anchor='middle' transform='rotate(-90 "
        f"{dim_x + text_offset:.2f} {depth / 2:.2f})'>"
        f"{_format_distance(depth)}</text>"
    )

    # Grid chain dimensions (horizontal)
    segments_x = list(zip(grid_x, grid_x[1:]))
    if segments_x:
        step = max(1, int(len(segments_x) / 8))
        y_chain = depth + grid_offset
        for idx, (x1, x2) in enumerate(segments_x):
            if idx % step != 0:
                continue
            dim_elements.append(
                f"<line x1='{x1:.2f}' y1='{y_chain:.2f}' x2='{x2:.2f}' y2='{y_chain:.2f}' "
                f"stroke='#a49384' stroke-width='{dim_weight:.2f}' />"
            )
            dim_elements.append(
                f"<line x1='{x1:.2f}' y1='{y_chain - tick * 0.6:.2f}' x2='{x1:.2f}' "
                f"y2='{y_chain + tick * 0.6:.2f}' stroke='#a49384' stroke-width='{dim_ext_weight:.2f}' />"
            )
            dim_elements.append(
                f"<line x1='{x2:.2f}' y1='{y_chain - tick * 0.6:.2f}' x2='{x2:.2f}' "
                f"y2='{y_chain + tick * 0.6:.2f}' stroke='#a49384' stroke-width='{dim_ext_weight:.2f}' />"
            )
            dim_elements.append(
                f"<text class='dim-text' x='{(x1 + x2) / 2:.2f}' y='{y_chain - text_offset * 0.6:.2f}' "
                "text-anchor='middle'>"
                f"{_format_distance(x2 - x1)}</text>"
            )

    # Grid chain dimensions (vertical)
    segments_y = list(zip(grid_y, grid_y[1:]))
    if segments_y:
        step = max(1, int(len(segments_y) / 8))
        x_chain = width + grid_offset
        for idx, (y1, y2) in enumerate(segments_y):
            if idx % step != 0:
                continue
            dim_elements.append(
                f"<line x1='{x_chain:.2f}' y1='{y1:.2f}' x2='{x_chain:.2f}' y2='{y2:.2f}' "
                f"stroke='#a49384' stroke-width='{dim_weight:.2f}' />"
            )
            dim_elements.append(
                f"<line x1='{x_chain - tick * 0.6:.2f}' y1='{y1:.2f}' "
                f"x2='{x_chain + tick * 0.6:.2f}' y2='{y1:.2f}' stroke='#a49384' stroke-width='{dim_ext_weight:.2f}' />"
            )
            dim_elements.append(
                f"<line x1='{x_chain - tick * 0.6:.2f}' y1='{y2:.2f}' "
                f"x2='{x_chain + tick * 0.6:.2f}' y2='{y2:.2f}' stroke='#a49384' stroke-width='{dim_ext_weight:.2f}' />"
            )
            dim_elements.append(
                f"<text class='dim-text' x='{x_chain + text_offset * 0.6:.2f}' y='{(y1 + y2) / 2:.2f}' "
                "text-anchor='middle' transform='rotate(-90 "
                f"{x_chain + text_offset * 0.6:.2f} {(y1 + y2) / 2:.2f})'>"
                f"{_format_distance(y2 - y1)}</text>"
            )

    # Architectural layer
    arch_elements.append(
        f"<rect x='0.5' y='0.5' width='{width - 1:.2f}' height='{depth - 1:.2f}' "
        f"fill='#fbf6ef' stroke='#c7b8a8' stroke-width='{wall_weight:.2f}' />"
    )
    arch_elements.append(
        f"<rect x='{core_x:.2f}' y='{core_y:.2f}' width='{core_w:.2f}' height='{core_d:.2f}' "
        f"fill='#e3d2c1' stroke='#b8a897' stroke-width='{partition_weight:.2f}' />"
    )
    arch_elements.append(
        f"<text class='space-label' x='{core_x + core_w / 2:.2f}' y='{core_y + core_d / 2:.2f}' "
        "text-anchor='middle'>CORE</text>"
    )

    for space in spaces:
        bounds = space.get("bounds") or {}
        min_x = _safe_number(bounds.get("min_x"), None)
        max_x = _safe_number(bounds.get("max_x"), None)
        min_y = _safe_number(bounds.get("min_y"), None)
        max_y = _safe_number(bounds.get("max_y"), None)
        if None in (min_x, max_x, min_y, max_y):
            continue
        min_x = _clamp(min_x, 0.5, width - 0.5)
        max_x = _clamp(max_x, 0.5, width - 0.5)
        min_y = _clamp(min_y, 0.5, depth - 0.5)
        max_y = _clamp(max_y, 0.5, depth - 0.5)
        width_s = max(max_x - min_x, 0.4)
        depth_s = max(max_y - min_y, 0.4)
        space_type = (space.get("type") or "space").lower()
        if space_type == "core":
            continue
        fill = palette.get(space_type, "#f7efe6")
        label = space_label(space_type, space.get("name"))
        arch_elements.append(
            f"<rect x='{min_x:.2f}' y='{min_y:.2f}' width='{width_s:.2f}' height='{depth_s:.2f}' "
            f"fill='{fill}' fill-opacity='0.85' stroke='#d7c7b7' stroke-width='{partition_weight:.2f}' />"
        )
        if space_type in {"apartment", "guest_room"}:
            orientation = entry_orientation(min_x, min_y, max_x, max_y)
            add_unit_detail(min_x, min_y, max_x, max_y, orientation)
        if width_s > 1.8 and depth_s > 1.8:
            arch_elements.append(
                f"<text class='space-label' x='{min_x + 0.6:.2f}' y='{min_y + 1.4:.2f}'>"
                f"{label}</text>"
            )
        if width_s > 2.8 and depth_s > 2.8:
            arch_elements.append(
                f"<text class='space-size' x='{min_x + 0.6:.2f}' y='{min_y + 2.3:.2f}'>"
                f"{width_s:.1f}x{depth_s:.1f}m</text>"
            )

        if space.get("requires_daylight"):
            if min_x <= 0.6:
                add_windows_vertical(min_y, max_y, 0.3, 1.1)
            if max_x >= width - 0.6:
                add_windows_vertical(min_y, max_y, width - 1.1, width - 0.3)
            if min_y <= 0.6:
                add_windows_horizontal(min_x, max_x, 0.3, 1.1)
            if max_y >= depth - 0.6:
                add_windows_horizontal(min_x, max_x, depth - 1.1, depth - 0.3)

        door_width = min(1.2, max(0.8, min(width_s, depth_s) * 0.22))
        if space_type not in {"core", "corridor", "circulation"}:
            orientation = entry_orientation(min_x, min_y, max_x, max_y)
            if orientation == "E":
                arch_elements.append(door_symbol(max_x, (min_y + max_y) / 2, door_width, "E"))
            elif orientation == "W":
                arch_elements.append(door_symbol(min_x, (min_y + max_y) / 2, door_width, "W"))
            elif orientation == "S":
                arch_elements.append(
                    door_symbol((min_x + max_x) / 2 - door_width / 2, max_y, door_width, "S")
                )
            else:
                arch_elements.append(
                    door_symbol((min_x + max_x) / 2 - door_width / 2, min_y, door_width, "N")
                )

    if not has_corridor_space:
        for corridor in circulation:
            path = corridor.get("path") or []
            if len(path) < 2:
                continue
            for idx in range(len(path) - 1):
                start = path[idx]
                end = path[idx + 1]
                arch_elements.append(
                    f"<line x1='{_safe_number(start.get('x'), 0):.2f}' y1='{_safe_number(start.get('y'), 0):.2f}' "
                    f"x2='{_safe_number(end.get('x'), 0):.2f}' y2='{_safe_number(end.get('y'), 0):.2f}' "
                    f"stroke='#c7b8a8' stroke-width='{detail_weight:.2f}' stroke-dasharray='2 2' />"
                )

    # Core symbols (elevators and stairs)
    if core_elevators > 0:
        elev_w = core_w * 0.18
        elev_d = core_d * 0.2
        elev_gap = elev_w * 0.2
        elev_x = core_x + elev_gap
        elev_y = core_y + elev_gap
        for idx in range(core_elevators):
            if elev_x + elev_w > core_x + core_w - elev_gap:
                break
            arch_elements.append(
                f"<rect x='{elev_x:.2f}' y='{elev_y:.2f}' width='{elev_w:.2f}' height='{elev_d:.2f}' "
                f"fill='#f3e9dd' stroke='#a08b77' stroke-width='{detail_weight:.2f}' />"
            )
            arch_elements.append(
                f"<text class='space-label' x='{elev_x + elev_w / 2:.2f}' "
                f"y='{elev_y + elev_d * 0.7:.2f}' text-anchor='middle'>E</text>"
            )
            elev_x += elev_w + elev_gap

    if core_stairs > 0:
        stair_w = core_w * 0.18
        stair_d = core_d * 0.35
        stair_gap = stair_w * 0.2
        stair_x = core_x + core_w - stair_w - stair_gap
        stair_y = core_y + core_d - stair_d - stair_gap
        for idx in range(core_stairs):
            if stair_y < core_y + stair_gap:
                break
            arch_elements.append(
                f"<rect x='{stair_x:.2f}' y='{stair_y:.2f}' width='{stair_w:.2f}' height='{stair_d:.2f}' "
                f"fill='#f3e9dd' stroke='#a08b77' stroke-width='{detail_weight:.2f}' />"
            )
            step_count = 5
            for step in range(step_count):
                y = stair_y + (step + 1) * stair_d / (step_count + 1)
                arch_elements.append(
                    f"<line x1='{stair_x:.2f}' y1='{y:.2f}' x2='{stair_x + stair_w:.2f}' "
                    f"y2='{y:.2f}' stroke='#9c8a78' stroke-width='{detail_weight:.2f}' />"
                )
            stair_y -= stair_d + stair_gap

    # Structural layer
    struct_grid = structural.get("grid", {}) if structural else {}
    struct_grid_x = _coerce_float_list(struct_grid.get("grid_x")) or grid_x
    struct_grid_y = _coerce_float_list(struct_grid.get("grid_y")) or grid_y
    columns = structural.get("columns") or []
    beams = [beam for beam in (structural.get("beams") or []) if beam.get("level", 0) == 0]

    if columns:
        for column in columns:
            pos = column.get("position") or {}
            col_x = _safe_number(pos.get("x"), 0)
            col_y = _safe_number(pos.get("y"), 0)
            col_w = _safe_number(column.get("width"), 0.4)
            col_d = _safe_number(column.get("depth"), 0.4)
            struct_elements.append(
                f"<rect x='{col_x - col_w / 2:.2f}' y='{col_y - col_d / 2:.2f}' "
                f"width='{col_w:.2f}' height='{col_d:.2f}' fill='#8a5a3c' "
                f"stroke='#6d4a34' stroke-width='{struct_light_weight:.2f}' />"
            )
    else:
        for x in struct_grid_x:
            for y in struct_grid_y:
                struct_elements.append(
                    f"<circle cx='{x:.2f}' cy='{y:.2f}' r='0.28' fill='#8a5a3c' />"
                )

    for beam in beams:
        start = beam.get("start") or {}
        end = beam.get("end") or {}
        struct_elements.append(
            f"<line x1='{_safe_number(start.get('x'), 0):.2f}' y1='{_safe_number(start.get('y'), 0):.2f}' "
            f"x2='{_safe_number(end.get('x'), 0):.2f}' y2='{_safe_number(end.get('y'), 0):.2f}' "
            f"stroke='#6f4f38' stroke-width='{struct_weight:.2f}' />"
        )

    # MEP layer
    hvac = mep.get("hvac", {}) if mep else {}
    electrical = mep.get("electrical", {}) if mep else {}
    plumbing = mep.get("plumbing", {}) if mep else {}
    fire_protection = mep.get("fire_protection", {}) if mep else {}

    duct_keys = set()
    for duct in hvac.get("ductwork", []) or []:
        start = duct.get("start") or {}
        end = duct.get("end") or {}
        x1 = _clamp(_safe_number(start.get("x"), 0), 0, width)
        y1 = _clamp(_safe_number(start.get("y"), 0), 0, depth)
        x2 = _clamp(_safe_number(end.get("x"), 0), 0, width)
        y2 = _clamp(_safe_number(end.get("y"), 0), 0, depth)
        if abs(x1 - x2) < 0.01 and abs(y1 - y2) < 0.01:
            continue
        key = tuple(sorted(((round(x1, 2), round(y1, 2)), (round(x2, 2), round(y2, 2)))))
        if key in duct_keys:
            continue
        duct_keys.add(key)
        duct_width = _safe_number(duct.get("width"), 0.6)
        stroke_width = max(mep_light_weight, min(mep_weight * 1.2, duct_width * 0.12))
        mep_elements.append(
            f"<line x1='{x1:.2f}' y1='{y1:.2f}' x2='{x2:.2f}' y2='{y2:.2f}' "
            f"stroke='#2d6f8e' stroke-width='{stroke_width:.2f}' />"
        )

    ahu_keys = set()
    for ahu in hvac.get("ahus", []) or []:
        loc = ahu.get("location") or {}
        dim = ahu.get("dimensions") or {}
        ahu_x = _safe_number(loc.get("x"), 0)
        ahu_y = _safe_number(loc.get("y"), 0)
        ahu_w = _safe_number(dim.get("l"), 1.2)
        ahu_d = _safe_number(dim.get("w"), 0.8)
        key = (round(ahu_x, 2), round(ahu_y, 2), round(ahu_w, 2), round(ahu_d, 2))
        if key in ahu_keys:
            continue
        ahu_keys.add(key)
        mep_elements.append(
            f"<rect x='{ahu_x:.2f}' y='{ahu_y:.2f}' width='{ahu_w:.2f}' height='{ahu_d:.2f}' "
            f"fill='#dceef4' stroke='#2d6f8e' stroke-width='{mep_light_weight:.2f}' />"
        )
        mep_elements.append(
            f"<text class='space-label' x='{ahu_x + ahu_w / 2:.2f}' y='{ahu_y + ahu_d / 1.4:.2f}' "
            "text-anchor='middle'>AHU</text>"
        )

    for riser in plumbing.get("risers", []) or []:
        pos = riser.get("position") or {}
        x = _safe_number(pos.get("x"), 0)
        y = _safe_number(pos.get("y"), 0)
        diameter = _safe_number(riser.get("diameter_mm"), 120)
        radius = max(0.35, min(0.7, diameter / 400))
        rtype = (riser.get("type") or "").lower()
        color = "#2b77c3"
        label = "R"
        if "hot" in rtype:
            color = "#d95c5c"
            label = "HW"
        elif "cold" in rtype:
            color = "#2b77c3"
            label = "CW"
        elif "drain" in rtype:
            color = "#4d9c6f"
            label = "DR"
        elif "vent" in rtype:
            color = "#8d8d8d"
            label = "V"
        mep_elements.append(
            f"<circle cx='{x:.2f}' cy='{y:.2f}' r='{radius:.2f}' fill='{color}' "
            f"stroke='#ffffff' stroke-width='{mep_light_weight:.2f}' />"
        )
        mep_elements.append(
            f"<text class='space-label' x='{x:.2f}' y='{y + radius + 0.6:.2f}' text-anchor='middle'>"
            f"{label}</text>"
        )

    distribution = electrical.get("distribution", {}) if electrical else {}
    for panel in distribution.get("panels", []) or []:
        loc = panel.get("location") or {}
        if _safe_number(loc.get("z"), 0) > 0.1:
            continue
        px = _safe_number(loc.get("x"), 0)
        py = _safe_number(loc.get("y"), 0)
        mep_elements.append(
            f"<rect x='{px - 0.6:.2f}' y='{py - 0.4:.2f}' width='1.2' height='0.8' "
            f"fill='#fde8d1' stroke='#d97927' stroke-width='{mep_light_weight:.2f}' />"
        )
        mep_elements.append(
            f"<text class='space-label' x='{px:.2f}' y='{py + 0.3:.2f}' text-anchor='middle'>"
            f"{panel.get('id', 'P')}</text>"
        )

    for route in distribution.get("cable_routes", []) or []:
        start = route.get("start") or {}
        end = route.get("end") or {}
        mep_elements.append(
            f"<line x1='{_safe_number(start.get('x'), 0):.2f}' y1='{_safe_number(start.get('y'), 0):.2f}' "
            f"x2='{_safe_number(end.get('x'), 0):.2f}' y2='{_safe_number(end.get('y'), 0):.2f}' "
            f"stroke='#d97927' stroke-width='{mep_light_weight:.2f}' stroke-dasharray='2 2' />"
        )

    diffusers = []
    for space in spaces:
        bounds = space.get("bounds") or {}
        min_x = _safe_number(bounds.get("min_x"), None)
        max_x = _safe_number(bounds.get("max_x"), None)
        min_y = _safe_number(bounds.get("min_y"), None)
        max_y = _safe_number(bounds.get("max_y"), None)
        if None in (min_x, max_x, min_y, max_y):
            continue
        space_type = (space.get("type") or "").lower()
        if space_type in {"corridor", "circulation", "core", "service"}:
            continue
        area = _safe_number(space.get("area"), (max_x - min_x) * (max_y - min_y))
        if area < 12:
            continue
        count = min(4, max(1, int(area / 60)))
        diffusers.extend(distribute_points(min_x, min_y, max_x, max_y, count))

    for x, y in diffusers[:40]:
        mep_elements.append(
            f"<path d='M{x - 0.3:.2f} {y + 0.3:.2f} L{x + 0.3:.2f} {y + 0.3:.2f} "
            f"L{x:.2f} {y - 0.3:.2f} Z' fill='#2d6f8e' />"
        )

    if fire_protection.get("sprinkler_system"):
        sprinkler_positions = []
        for i in range(len(grid_x) - 1):
            for j in range(len(grid_y) - 1):
                if len(sprinkler_positions) >= 24:
                    break
                sprinkler_positions.append(
                    ((grid_x[i] + grid_x[i + 1]) / 2, (grid_y[j] + grid_y[j + 1]) / 2)
                )
        for x, y in sprinkler_positions:
            mep_elements.append(
                f"<line x1='{x - 0.3:.2f}' y1='{y:.2f}' x2='{x + 0.3:.2f}' y2='{y:.2f}' "
                f"stroke='#c0392b' stroke-width='{mep_light_weight:.2f}' />"
            )
            mep_elements.append(
                f"<line x1='{x:.2f}' y1='{y - 0.3:.2f}' x2='{x:.2f}' y2='{y + 0.3:.2f}' "
                f"stroke='#c0392b' stroke-width='{mep_light_weight:.2f}' />"
            )

    # Legend + north arrow + scale bar
    legend_x = -margin + 0.6
    legend_y = -margin + 0.6
    legend_items = [
        ("Walls", "line", "#c7b8a8"),
        ("Door", "door", "#7d6a58"),
        ("Window", "window", "#8aa6c1"),
        ("Column", "circle", "#8a5a3c"),
        ("Beam", "line", "#6f4f38"),
        ("Duct", "line", "#2d6f8e"),
        ("Riser", "circle", "#2b77c3"),
        ("Panel", "rect", "#d97927"),
    ]
    row_h = max(1.3, margin * 0.22)
    legend_w = max(14.0, margin * 2.6)
    legend_h = row_h * (len(legend_items) + 2.2)
    legend_elements.append(
        f"<rect x='{legend_x:.2f}' y='{legend_y:.2f}' width='{legend_w:.2f}' height='{legend_h:.2f}' "
        f"fill='#fff7ef' fill-opacity='0.95' stroke='#c7b8a8' stroke-width='{legend_weight:.2f}' rx='0.6' />"
    )
    legend_elements.append(
        f"<text class='legend-text' x='{legend_x + 0.8:.2f}' y='{legend_y + row_h:.2f}'>"
        "Coordination Legend</text>"
    )
    for idx, (label, icon, color) in enumerate(legend_items):
        y = legend_y + row_h * (idx + 1.8)
        icon_x = legend_x + 0.8
        if icon == "line":
            legend_elements.append(
                f"<line x1='{icon_x:.2f}' y1='{y - 0.3:.2f}' x2='{icon_x + 2.0:.2f}' y2='{y - 0.3:.2f}' "
                f"stroke='{color}' stroke-width='{legend_weight:.2f}' />"
            )
        elif icon == "door":
            legend_elements.append(door_symbol(icon_x, y - 0.3, 1.2, "N"))
        elif icon == "window":
            legend_elements.append(
                f"<line x1='{icon_x:.2f}' y1='{y - 0.2:.2f}' x2='{icon_x + 1.6:.2f}' y2='{y - 0.2:.2f}' "
                f"stroke='{color}' stroke-width='{legend_weight:.2f}' />"
            )
            legend_elements.append(
                f"<line x1='{icon_x:.2f}' y1='{y + 0.1:.2f}' x2='{icon_x + 1.6:.2f}' y2='{y + 0.1:.2f}' "
                f"stroke='{color}' stroke-width='{legend_weight:.2f}' />"
            )
        elif icon == "circle":
            legend_elements.append(
                f"<circle cx='{icon_x + 0.9:.2f}' cy='{y - 0.2:.2f}' r='0.35' fill='{color}' />"
            )
        elif icon == "rect":
            legend_elements.append(
                f"<rect x='{icon_x:.2f}' y='{y - 0.6:.2f}' width='1.6' height='0.9' "
                f"fill='{color}' stroke='#8a5a3c' stroke-width='{legend_weight:.2f}' />"
            )
        legend_elements.append(
            f"<text class='legend-text' x='{icon_x + 3.0:.2f}' y='{y:.2f}'>"
            f"{label}</text>"
        )

    # North arrow
    north_x = legend_x + legend_w - 2.5
    north_y = legend_y + row_h * 0.9
    legend_elements.append(
        f"<path d='M{north_x:.2f} {north_y + 1.6:.2f} L{north_x + 0.8:.2f} {north_y:.2f} "
        f"L{north_x + 1.6:.2f} {north_y + 1.6:.2f} Z' fill='#6f4f38' />"
    )
    legend_elements.append(
        f"<text class='legend-text' x='{north_x + 0.8:.2f}' y='{north_y + 2.5:.2f}' text-anchor='middle'>N</text>"
    )

    # Scale bar
    scale_len = min(10.0, max(6.0, width * 0.18))
    scale_x = legend_x + 0.8
    scale_y = legend_y + legend_h - row_h * 0.6
    legend_elements.append(
        f"<line x1='{scale_x:.2f}' y1='{scale_y:.2f}' x2='{scale_x + scale_len:.2f}' y2='{scale_y:.2f}' "
        f"stroke='#6f4f38' stroke-width='{legend_weight:.2f}' />"
    )
    legend_elements.append(
        f"<line x1='{scale_x:.2f}' y1='{scale_y - 0.4:.2f}' x2='{scale_x:.2f}' y2='{scale_y + 0.4:.2f}' "
        f"stroke='#6f4f38' stroke-width='{legend_weight:.2f}' />"
    )
    legend_elements.append(
        f"<line x1='{scale_x + scale_len / 2:.2f}' y1='{scale_y - 0.3:.2f}' "
        f"x2='{scale_x + scale_len / 2:.2f}' y2='{scale_y + 0.3:.2f}' stroke='#6f4f38' stroke-width='{detail_weight:.2f}' />"
    )
    legend_elements.append(
        f"<line x1='{scale_x + scale_len:.2f}' y1='{scale_y - 0.4:.2f}' "
        f"x2='{scale_x + scale_len:.2f}' y2='{scale_y + 0.4:.2f}' stroke='#6f4f38' stroke-width='{legend_weight:.2f}' />"
    )
    legend_elements.append(
        f"<text class='legend-text' x='{scale_x:.2f}' y='{scale_y + 1.2:.2f}'>0</text>"
    )
    legend_elements.append(
        f"<text class='legend-text' x='{scale_x + scale_len / 2:.2f}' y='{scale_y + 1.2:.2f}' "
        "text-anchor='middle'>"
        f"{_format_distance(scale_len / 2).replace(' m', '')}</text>"
    )
    legend_elements.append(
        f"<text class='legend-text' x='{scale_x + scale_len:.2f}' y='{scale_y + 1.2:.2f}' "
        "text-anchor='end'>"
        f"{_format_distance(scale_len).replace(' m', '')}m</text>"
    )

    elements.append("<g class='layer-axes'>" + "".join(axis_elements) + "</g>")
    elements.append("<g class='layer-dims'>" + "".join(dim_elements) + "</g>")
    elements.append("<g class='layer-arch'>" + "".join(arch_elements) + "</g>")
    if struct_elements:
        elements.append("<g class='layer-struct'>" + "".join(struct_elements) + "</g>")
    if mep_elements:
        elements.append("<g class='layer-mep'>" + "".join(mep_elements) + "</g>")
    elements.append("<g class='layer-legend'>" + "".join(legend_elements) + "</g>")
    elements.append("</svg>")
    return "".join(elements)


def _build_structural_plan_svg(structural: dict, massing: dict) -> str:
    width = _safe_number(massing.get("width"), 40)
    depth = _safe_number(massing.get("depth"), 30)
    grid = structural.get("grid", {})
    grid_x = _coerce_float_list(grid.get("grid_x"))
    grid_y = _coerce_float_list(grid.get("grid_y"))
    module = _safe_number(massing.get("module"), 6)

    if not grid_x:
        count = max(int(width / module), 2)
        grid_x = [i * width / count for i in range(count + 1)]
    if not grid_y:
        count = max(int(depth / module), 2)
        grid_y = [i * depth / count for i in range(count + 1)]

    columns = structural.get("columns") or []
    beams = [beam for beam in (structural.get("beams") or []) if beam.get("level", 0) == 0]
    wall_weight = 1.2
    grid_weight = 0.45
    beam_weight = 0.8
    column_weight = 0.75

    elements = [
        f"<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 {width:.2f} {depth:.2f}' "
        f"data-grid-x='{','.join(f'{value:.2f}' for value in grid_x)}' "
        f"data-grid-y='{','.join(f'{value:.2f}' for value in grid_y)}' "
        "shape-rendering='geometricPrecision'>",
        "<style>[stroke]{vector-effect:non-scaling-stroke;stroke-linecap:square;"
        "stroke-linejoin:miter;}</style>",
        f"<rect x='0.5' y='0.5' width='{width - 1:.2f}' height='{depth - 1:.2f}' "
        f"fill='#fbf6ef' stroke='#c7b8a8' stroke-width='{wall_weight:.2f}' />",
    ]

    for x in grid_x:
        elements.append(
            f"<line x1='{x:.2f}' y1='0.5' x2='{x:.2f}' y2='{depth - 0.5:.2f}' "
            f"stroke='#e3d5c6' stroke-width='{grid_weight:.2f}' />"
        )
    for y in grid_y:
        elements.append(
            f"<line x1='0.5' y1='{y:.2f}' x2='{width - 0.5:.2f}' y2='{y:.2f}' "
            f"stroke='#e3d5c6' stroke-width='{grid_weight:.2f}' />"
        )

    for beam in beams:
        start = beam.get("start") or {}
        end = beam.get("end") or {}
        elements.append(
            f"<line x1='{_safe_number(start.get('x'), 0):.2f}' y1='{_safe_number(start.get('y'), 0):.2f}' "
            f"x2='{_safe_number(end.get('x'), 0):.2f}' y2='{_safe_number(end.get('y'), 0):.2f}' "
            f"stroke='#9c7b5d' stroke-width='{beam_weight:.2f}' />"
        )

    for column in columns:
        pos = column.get("position") or {}
        col_x = _safe_number(pos.get("x"), 0)
        col_y = _safe_number(pos.get("y"), 0)
        col_w = _safe_number(column.get("width"), 0.4)
        col_d = _safe_number(column.get("depth"), 0.4)
        elements.append(
            f"<rect x='{col_x - col_w / 2:.2f}' y='{col_y - col_d / 2:.2f}' "
            f"width='{col_w:.2f}' height='{col_d:.2f}' fill='#b98b6c' "
            f"stroke='#83624b' stroke-width='{column_weight:.2f}' />"
        )

    elements.append("</svg>")
    return "".join(elements)


def _build_ifc_stub(project: models.Project) -> str:
    name = project.name or "AI Designer Project"
    return (
        "ISO-10303-21;\n"
        "HEADER;\n"
        "FILE_DESCRIPTION(('ViewDefinition [CoordinationView]'), '2;1');\n"
        "FILE_NAME('AIDesigner.ifc','2025-01-01T00:00:00',('AI Designer'),('AI Designer'),'', '', '');\n"
        "FILE_SCHEMA(('IFC4'));\n"
        "ENDSEC;\n"
        "DATA;\n"
        f"#1=IFCPROJECT('0hY3YQnS92X9QkK7CzA1gG',$,'{name}',$,$,$,$,(#2),#3);\n"
        "#2=IFCGEOMETRICREPRESENTATIONCONTEXT($,'Model',3,0.0001,#4,$);\n"
        "#3=IFCUNITASSIGNMENT((#5,#6,#7));\n"
        "#4=IFCAXIS2PLACEMENT3D(#8,#9,#10);\n"
        "#5=IFCSIUNIT(*,.LENGTHUNIT.,$,.METRE.);\n"
        "#6=IFCSIUNIT(*,.AREAUNIT.,$,.SQUARE_METRE.);\n"
        "#7=IFCSIUNIT(*,.VOLUMEUNIT.,$,.CUBIC_METRE.);\n"
        "#8=IFCCARTESIANPOINT((0.,0.,0.));\n"
        "#9=IFCDIRECTION((0.,0.,1.));\n"
        "#10=IFCDIRECTION((1.,0.,0.));\n"
        "ENDSEC;\n"
        "END-ISO-10303-21;\n"
    )


def _build_gltf_stub(project: models.Project) -> str:
    massing = _compute_massing(project)
    width = massing["width"]
    depth = massing["depth"]
    height = massing["height"]

    positions = [
        -0.5,
        0.0,
        -0.5,
        0.5,
        0.0,
        -0.5,
        0.5,
        0.0,
        0.5,
        -0.5,
        0.0,
        0.5,
        -0.5,
        1.0,
        -0.5,
        0.5,
        1.0,
        -0.5,
        0.5,
        1.0,
        0.5,
        -0.5,
        1.0,
        0.5,
    ]

    indices = [
        0,
        1,
        2,
        2,
        3,
        0,
        4,
        5,
        6,
        6,
        7,
        4,
        0,
        1,
        5,
        5,
        4,
        0,
        1,
        2,
        6,
        6,
        5,
        1,
        2,
        3,
        7,
        7,
        6,
        2,
        3,
        0,
        4,
        4,
        7,
        3,
    ]

    pos_bytes = struct.pack("<" + "f" * len(positions), *positions)
    idx_bytes = struct.pack("<" + "H" * len(indices), *indices)
    buffer = pos_bytes + idx_bytes
    encoded = base64.b64encode(buffer).decode("ascii")

    return f"""{{
  "asset": {{ "version": "2.0", "generator": "AI Designer" }},
  "scene": 0,
  "scenes": [{{ "nodes": [0] }}],
  "nodes": [{{ "mesh": 0, "scale": [{width:.2f}, {height:.2f}, {depth:.2f}] }}],
  "meshes": [{{
    "primitives": [{{
      "attributes": {{ "POSITION": 0 }},
      "indices": 1
    }}]
  }}],
  "buffers": [{{
    "uri": "data:application/octet-stream;base64,{encoded}",
    "byteLength": {len(buffer)}
  }}],
  "bufferViews": [
    {{ "buffer": 0, "byteOffset": 0, "byteLength": {len(pos_bytes)} }},
    {{ "buffer": 0, "byteOffset": {len(pos_bytes)}, "byteLength": {len(idx_bytes)} }}
  ],
  "accessors": [
    {{
      "bufferView": 0,
      "componentType": 5126,
      "count": 8,
      "type": "VEC3",
      "min": [-0.5, 0.0, -0.5],
      "max": [0.5, 1.0, 0.5]
    }},
    {{
      "bufferView": 1,
      "componentType": 5123,
      "count": {len(indices)},
      "type": "SCALAR"
    }}
  ]
}}"""


def _build_dxf(
    plan: dict,
    structural: Optional[dict] = None,
    mep: Optional[dict] = None,
    massing: Optional[dict] = None,
):
    """Convert plan data to DXF format for CAD software compatibility."""
    if ezdxf is None:
        return None

    plan = plan or {}
    structural = structural or {}
    mep = mep or {}
    massing = massing or {}

    doc = ezdxf.new("R2018")  # AutoCAD 2018 format
    doc.units = dxf_units.M  # Meters

    msp = doc.modelspace()

    # Professional CAD line weights (in hundredths of mm)
    # Matching AutoCAD/Revit standards for architectural drawings
    LW_WALL = 70        # 0.70mm - Walls/building outline (thick)
    LW_PARTITION = 50   # 0.50mm - Partitions/core (medium)
    LW_DOOR = 35        # 0.35mm - Doors (thin)
    LW_WINDOW = 35      # 0.35mm - Windows (thin)
    LW_COLUMN = 50      # 0.50mm - Structural columns (medium)
    LW_GRID = 18        # 0.18mm - Grid lines (hairline)
    LW_DIM = 25         # 0.25mm - Dimensions (thin)
    LW_SPACE = 25       # 0.25mm - Space outlines (thin)
    LW_MEP = 35         # 0.35mm - MEP elements (thin)

    # Create layers with standard CAD colors and line weights
    doc.layers.add("WALLS", color=7, lineweight=LW_WALL)  # White
    doc.layers.add("DOORS", color=30, lineweight=LW_DOOR)  # Orange
    doc.layers.add("WINDOWS", color=140, lineweight=LW_WINDOW)  # Blue
    doc.layers.add("COLUMNS", color=1, lineweight=LW_COLUMN)  # Red
    doc.layers.add("GRID", color=8, lineweight=LW_GRID)  # Gray
    doc.layers.add("DIMENSIONS", color=2, lineweight=LW_DIM)  # Yellow
    doc.layers.add("CORE", color=252, lineweight=LW_PARTITION)  # Light gray
    doc.layers.add("SPACES", color=254, lineweight=LW_SPACE)  # Very light gray
    doc.layers.add("MEP-HVAC", color=4, lineweight=LW_MEP)  # Cyan
    doc.layers.add("MEP-ELECTRICAL", color=6, lineweight=LW_MEP)  # Magenta
    doc.layers.add("MEP-PLUMBING", color=5, lineweight=LW_MEP)  # Blue

    # Extract dimensions
    width = _safe_number(plan.get("width"), _safe_number(massing.get("width"), 40))
    depth = _safe_number(plan.get("depth"), _safe_number(massing.get("depth"), 30))

    # Draw building outline with professional wall weight
    msp.add_lwpolyline(
        [(0, 0), (width, 0), (width, depth), (0, depth)],
        close=True,
        dxfattribs={"layer": "WALLS", "lineweight": LW_WALL},
    )

    # Draw grid lines
    grid_x = _coerce_float_list(plan.get("grid_x")) or _coerce_float_list(massing.get("grid_x"))
    grid_y = _coerce_float_list(plan.get("grid_y")) or _coerce_float_list(massing.get("grid_y"))

    # Generate default grid if none provided
    if not grid_x:
        grid_count = int(_safe_number(massing.get("grid_x"), 4))
        grid_x = [i * width / max(grid_count - 1, 1) for i in range(grid_count)]
    if not grid_y:
        grid_count = int(_safe_number(massing.get("grid_y"), 4))
        grid_y = [i * depth / max(grid_count - 1, 1) for i in range(grid_count)]

    for x in grid_x:
        msp.add_line((x, 0), (x, depth), dxfattribs={"layer": "GRID", "lineweight": LW_GRID})
    for y in grid_y:
        msp.add_line((0, y), (width, y), dxfattribs={"layer": "GRID", "lineweight": LW_GRID})

    # Draw core
    core = plan.get("core") or {}
    if core:
        core_pos = core.get("position") or [width * 0.35, depth * 0.35]
        try:
            core_x = float(core_pos[0])
            core_y = float(core_pos[1])
        except (TypeError, IndexError, ValueError):
            core_x = width * 0.35
            core_y = depth * 0.35
        core_w = _safe_number(core.get("width"), width * 0.3)
        core_d = _safe_number(core.get("depth"), depth * 0.3)
    else:
        core_ratio = _safe_number(massing.get("core_ratio"), 0.2)
        core_w = width * 0.32
        core_area = width * depth * core_ratio
        core_d = min(core_area / max(core_w, 1), depth * 0.6)
        core_x = (width - core_w) / 2
        core_y = (depth - core_d) / 2

    msp.add_lwpolyline(
        [
            (core_x, core_y),
            (core_x + core_w, core_y),
            (core_x + core_w, core_y + core_d),
            (core_x, core_y + core_d),
        ],
        close=True,
        dxfattribs={"layer": "CORE"},
    )
    msp.add_text(
        "CORE",
        dxfattribs={
            "layer": "DIMENSIONS",
            "height": min(core_w, core_d) * 0.15,
            "insert": (core_x + core_w / 2, core_y + core_d / 2),
        },
    )

    # Draw columns from structural data
    columns = structural.get("columns") or []
    if columns:
        for col in columns:
            pos = col.get("position") or {}
            cx = _safe_number(pos.get("x"), 0)
            cy = _safe_number(pos.get("y"), 0)
            col_w = _safe_number(col.get("width"), 0.4)
            col_d = _safe_number(col.get("depth"), 0.4)
            msp.add_lwpolyline(
                [
                    (cx - col_w / 2, cy - col_d / 2),
                    (cx + col_w / 2, cy - col_d / 2),
                    (cx + col_w / 2, cy + col_d / 2),
                    (cx - col_w / 2, cy + col_d / 2),
                ],
                close=True,
                dxfattribs={"layer": "COLUMNS"},
            )
    else:
        # Draw columns at grid intersections
        for x in grid_x:
            for y in grid_y:
                msp.add_circle((x, y), 0.2, dxfattribs={"layer": "COLUMNS"})

    # Draw spaces
    for space in plan.get("spaces") or []:
        bounds = space.get("bounds") or {}
        min_x = _safe_number(bounds.get("min_x"), None)
        max_x = _safe_number(bounds.get("max_x"), None)
        min_y = _safe_number(bounds.get("min_y"), None)
        max_y = _safe_number(bounds.get("max_y"), None)

        if None in (min_x, max_x, min_y, max_y):
            continue

        space_type = (space.get("type") or "space").lower()
        if space_type == "core":
            continue

        msp.add_lwpolyline(
            [(min_x, min_y), (max_x, min_y), (max_x, max_y), (min_x, max_y)],
            close=True,
            dxfattribs={"layer": "SPACES"},
        )

        # Add space label
        center_x = (min_x + max_x) / 2
        center_y = (min_y + max_y) / 2
        space_name = space.get("name") or space_type.replace("_", " ").title()
        space_w = max_x - min_x
        space_d = max_y - min_y

        if space_w > 2 and space_d > 2:
            msp.add_text(
                space_name[:15],
                dxfattribs={
                    "layer": "DIMENSIONS",
                    "height": min(space_w, space_d) * 0.08,
                    "insert": (center_x, center_y),
                },
            )

    # Draw MEP elements
    hvac = mep.get("hvac") or {}
    electrical = mep.get("electrical") or {}
    plumbing = mep.get("plumbing") or {}

    # HVAC ductwork
    for duct in hvac.get("ductwork") or []:
        start = duct.get("start") or {}
        end = duct.get("end") or {}
        x1 = _safe_number(start.get("x"), 0)
        y1 = _safe_number(start.get("y"), 0)
        x2 = _safe_number(end.get("x"), 0)
        y2 = _safe_number(end.get("y"), 0)
        if abs(x1 - x2) > 0.01 or abs(y1 - y2) > 0.01:
            msp.add_line((x1, y1), (x2, y2), dxfattribs={"layer": "MEP-HVAC"})

    # Electrical panels
    for panel in (electrical.get("distribution") or {}).get("panels") or []:
        loc = panel.get("location") or {}
        px = _safe_number(loc.get("x"), 0)
        py = _safe_number(loc.get("y"), 0)
        if _safe_number(loc.get("z"), 0) < 0.1:  # Only ground floor
            msp.add_circle((px, py), 0.3, dxfattribs={"layer": "MEP-ELECTRICAL"})
            msp.add_text(
                panel.get("id", "P"),
                dxfattribs={
                    "layer": "MEP-ELECTRICAL",
                    "height": 0.2,
                    "insert": (px + 0.4, py),
                },
            )

    # Plumbing risers
    for riser in plumbing.get("risers") or []:
        pos = riser.get("position") or {}
        rx = _safe_number(pos.get("x"), 0)
        ry = _safe_number(pos.get("y"), 0)
        msp.add_circle((rx, ry), 0.15, dxfattribs={"layer": "MEP-PLUMBING"})

    # Add dimensions
    try:
        dim_style = doc.dimstyles.new("AI_DESIGNER")
        dim_style.dxf.dimtxt = 0.3
        dim_style.dxf.dimasz = 0.2

        # Width dimension
        msp.add_linear_dim(
            base=(width / 2, -2),
            p1=(0, 0),
            p2=(width, 0),
            dimstyle="AI_DESIGNER",
            dxfattribs={"layer": "DIMENSIONS"},
        ).render()

        # Depth dimension
        msp.add_linear_dim(
            base=(width + 2, depth / 2),
            p1=(width, 0),
            p2=(width, depth),
            angle=90,
            dimstyle="AI_DESIGNER",
            dxfattribs={"layer": "DIMENSIONS"},
        ).render()
    except Exception:
        # Dimensions are optional, don't fail if they can't be added
        pass

    return doc


def _write_dxf(path: str, doc) -> bool:
    """Write DXF file to disk."""
    if doc is None or ezdxf is None:
        return False
    try:
        doc.saveas(path)
        return True
    except Exception:
        return False


def _write_schedule(
    path: str,
    project: models.Project,
    massing: dict,
    structural: Optional[dict] = None,
    mep: Optional[dict] = None,
) -> None:
    lines = [
        "Section,Item,Value",
        f"Summary,Project,{project.name}",
        f"Summary,Region,{project.region}",
        f"Summary,Building Type,{project.building_type}",
        f"Summary,GFA (m2),{_safe_number(project.gfa, massing['floor_area'] * massing['floors']):.0f}",
        f"Summary,Floors,{massing['floors']:.0f}",
    ]

    if Workbook is None:
        _write_file(path, "\n".join(lines))
        return

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Section", "Item", "Value"])
    summary.append(["Summary", "Project", project.name])
    summary.append(["Summary", "Region", project.region])
    summary.append(["Summary", "Building Type", project.building_type])
    summary.append(["Summary", "Floor Plate (m2)", f"{massing['floor_area']:.0f}"])
    summary.append(
        [
            "Summary",
            "GFA (m2)",
            f"{_safe_number(project.gfa, massing['floor_area'] * massing['floors']):.0f}",
        ]
    )
    summary.append(["Summary", "Floors", f"{massing['floors']:.0f}"])

    structure = workbook.create_sheet("Structure")
    grid_x = massing["grid_x"]
    grid_y = massing["grid_y"]
    column_count = grid_x * grid_y * int(max(massing["floors"], 1))
    structure.append(["Item", "Value"])
    structure.append(["Grid X", grid_x])
    structure.append(["Grid Y", grid_y])
    structure.append(["Module (m)", f"{massing['module']:.1f}"])
    structure.append(["Estimated Columns", column_count])
    if structural:
        structure.append(["System", structural.get("system", "")])
        structure.append(["Material", structural.get("material", "")])
        metrics = structural.get("metrics", {})
        structure.append(["Avg Column Utilization", f"{metrics.get('avg_column_utilization', 0):.2f}"])
        structure.append(["Avg Beam Utilization", f"{metrics.get('avg_beam_utilization', 0):.2f}"])
        structure.append(["Max Drift Ratio", f"{metrics.get('max_drift_ratio', 0):.4f}"])

        member_sheet = workbook.create_sheet("Structural Elements")
        member_sheet.append(["Type", "ID", "Width (m)", "Depth (m)", "Utilization"])
        for col in (structural.get("columns") or [])[:250]:
            member_sheet.append(
                [
                    "Column",
                    col.get("id"),
                    f"{_safe_number(col.get('width'), 0):.2f}",
                    f"{_safe_number(col.get('depth'), 0):.2f}",
                    f"{_safe_number(col.get('utilization'), 0):.2f}",
                ]
            )
        for beam in (structural.get("beams") or [])[:250]:
            member_sheet.append(
                [
                    "Beam",
                    beam.get("id"),
                    f"{_safe_number(beam.get('width'), 0):.2f}",
                    f"{_safe_number(beam.get('depth'), 0):.2f}",
                    f"{_safe_number(beam.get('utilization'), 0):.2f}",
                ]
            )

    mep_sheet = workbook.create_sheet("MEP")
    mep_sheet.append(["Item", "Value"])
    mep_sheet.append(["Risers", 4])
    mep_sheet.append(["HVAC Zones", max(int(massing["floors"] / 3), 3)])
    mep_sheet.append(["Electrical Panels", max(int(massing["floors"] / 4), 2)])
    if mep:
        hvac = mep.get("hvac", {})
        electrical = mep.get("electrical", {})
        plumbing = mep.get("plumbing", {})
        mep_sheet.append(["HVAC System", hvac.get("system_type", "")])
        mep_sheet.append(
            ["Cooling Capacity (kW)", f"{_safe_number(hvac.get('cooling_capacity_kw'), 0):.1f}"]
        )
        mep_sheet.append(
            [
                "Total Demand (kVA)",
                f"{_safe_number(electrical.get('loads', {}).get('total_demand'), 0):.0f}",
            ]
        )
        mep_sheet.append(["Riser Count", len(plumbing.get("risers", []))])

    workbook.save(path)


def _pdf_escape(text: str) -> str:
    safe_text = text.encode("latin-1", "ignore").decode("latin-1")
    return safe_text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _build_pdf_report(lines: List[str]) -> bytes:
    content_lines = ["BT", "/F1 12 Tf", "72 740 Td"]
    for line in lines:
        content_lines.append(f"({_pdf_escape(line)}) Tj")
        content_lines.append("0 -14 Td")
    content_lines.append("ET")
    content = "\n".join(content_lines)
    content_bytes = content.encode("latin-1")

    objects = []
    objects.append("1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n")
    objects.append("2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n")
    objects.append(
        "3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        "/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>endobj\n"
    )
    objects.append(
        f"4 0 obj<< /Length {len(content_bytes)} >>stream\n{content}\nendstream\nendobj\n"
    )
    objects.append("5 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n")

    pdf = "%PDF-1.4\n"
    offsets = []
    for obj in objects:
        offsets.append(len(pdf))
        pdf += obj

    xref_start = len(pdf)
    pdf += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
    for offset in offsets:
        pdf += f"{offset:010d} 00000 n \n"
    pdf += (
        f"trailer<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_start}\n%%EOF\n"
    )

    return pdf.encode("latin-1")


def _write_review_package(run_dir: str, package_name: str, files: list[str]) -> None:
    package_path = os.path.join(run_dir, package_name)
    with zipfile.ZipFile(package_path, "w") as archive:
        for file_name in files:
            file_path = os.path.join(run_dir, file_name)
            if os.path.exists(file_path):
                archive.write(file_path, arcname=file_name)


def apply_state_payload(
    db: Session, payload
) -> Tuple[models.Project, Optional[models.Run], Optional[models.Output]]:
    project = None
    if payload.project.id:
        project = db.query(models.Project).filter(models.Project.id == payload.project.id).first()
    if not project:
        project = models.Project(name=payload.project.name)
        db.add(project)

    project.name = payload.project.name
    project.region = _safe_text(payload.project.region)
    project.building_type = _safe_text(payload.project.type)
    project.phase = _safe_text(payload.project.phase)
    project.gfa = _safe_text(payload.project.gfa)
    project.floors = _safe_text(payload.project.floors)
    project.status = _safe_text(payload.project.status)
    project.next_run = _safe_text(payload.project.next_run)
    project.brief = _safe_text(payload.project.brief)
    project.core_ratio = _safe_text(payload.project.core_ratio)
    project.parking = _safe_text(payload.project.parking)
    project.budget = _safe_text(payload.project.budget)
    project.delivery = _safe_text(payload.project.delivery)
    project.code_library = _safe_text(payload.project.code_library)
    project.energy_target = _safe_text(payload.project.energy_target)
    project.daylight = _safe_text(payload.project.daylight)
    project.structural_system = _safe_text(payload.project.structural_system)
    project.mep_strategy = _safe_text(payload.project.mep_strategy)
    project.site_model = _safe_text(payload.project.site_model)

    run = None
    if payload.run:
        if payload.run.id:
            run = db.query(models.Run).filter(models.Run.id == payload.run.id).first()
        if not run:
            run = models.Run(project=project)
            db.add(run)
        run.status = _safe_text(payload.run.status)
        run.conflicts = _safe_text(payload.run.conflicts)

    output = None
    if payload.outputs:
        if payload.outputs.id:
            output = db.query(models.Output).filter(models.Output.id == payload.outputs.id).first()
        if not output and run:
            output = models.Output(run=run)
            db.add(output)
        if output:
            output.clash_density = _safe_text(payload.outputs.clash_density)
            output.structural_variance = _safe_text(payload.outputs.structural_variance)
            output.compliance = _safe_text(payload.outputs.compliance)
            output.energy = _safe_text(payload.outputs.energy)
            output.clash_free = payload.outputs.clash_free or 0
            output.energy_score = payload.outputs.energy_score or 0
            output.structural_score = payload.outputs.structural_score or 0
            output.ifc_file = _safe_text(payload.outputs.ifc_file)
            output.mep_schedule_file = _safe_text(payload.outputs.mep_schedule_file)
            output.energy_report_file = _safe_text(payload.outputs.energy_report_file)
            output.review_package_file = _safe_text(payload.outputs.review_package_file)
            output.plan_svg_file = _safe_text(payload.outputs.plan_svg_file)
            output.gltf_file = _safe_text(payload.outputs.gltf_file)
            output.generated_at = _safe_text(payload.outputs.generated_at)

    db.commit()
    db.refresh(project)
    if run:
        db.refresh(run)
    if output:
        db.refresh(output)
    return project, run, output


def get_latest_state(
    db: Session,
) -> Tuple[Optional[models.Project], Optional[models.Run], Optional[models.Output]]:
    project = db.query(models.Project).order_by(models.Project.created_at.desc()).first()
    if not project:
        return None, None, None
    run = (
        db.query(models.Run)
        .filter(models.Run.project_id == project.id)
        .order_by(models.Run.created_at.desc())
        .first()
    )
    output = None
    if run:
        output = db.query(models.Output).filter(models.Output.run_id == run.id).first()
    return project, run, output


def start_run(
    db: Session, project_id: Optional[str]
) -> Tuple[models.Project, models.Run, models.Output]:
    if project_id:
        project = db.query(models.Project).filter(models.Project.id == project_id).first()
    else:
        project = db.query(models.Project).order_by(models.Project.created_at.desc()).first()

    if not project:
        raise ValueError("Project not found")

    run = models.Run(project=project, status="In progress", conflicts="Scanning")
    run.started_at = datetime.utcnow()
    db.add(run)
    db.flush()

    output = models.Output(run=run)
    db.add(output)
    db.flush()

    project.status = "Running"
    project.next_run = "Orchestrator active"

    db.commit()
    db.refresh(project)
    db.refresh(run)
    db.refresh(output)

    return project, run, output
